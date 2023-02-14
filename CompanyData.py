import openpyxl
import OpenBrowser
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import re

class CompanyDatas():
    def datas(self):
        CompanyUrl = openpyxl.load_workbook("Result/Companies.xlsx")
        driver.implicitly_wait(2)
        pattern = r'\d+\.?\d*'
        sn = CompanyUrl.get_sheet_by_name('Sheet')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "SN"
        ws['B1'] = "Name"
        ws['C1'] = "Trading Code"
        ws['D1'] = "Opening Price"
        ws['E1'] = "Last Day Closing Price"
        ws['F1'] = "Last Trade Price"
        ws['G1'] = "Closing Price"
        ws['H1'] = "Volume"
        ws['I1'] = "Trade"
        ws['J1'] = "Category"
        ws['K1'] = "Cash Dividend"
        ws['L1'] = "Stock Dividend"
        ws['M1'] = "Debaut Trading Code"
        ws['N1'] = "SHP-Sponsor/director"
        ws['O1'] = "SHP-Govt"
        ws['P1'] = "SHP-Institute"
        ws['Q1'] = "SHP-Foreign"
        ws['R1'] = "SHP-Public"
        ws['S1'] = "Company Address"
        ws['T1'] = "Contact"
        ws['U1'] = "Email"
        ws['V1'] = "Website"
        ws['W1'] = "DSE URL"
        for x in range(2, 1000):
            CUrl = sn['B' + str(x)].value
            CurlSn = sn['A' + str(x)].value
            if(CUrl != None):
                driver.get(CUrl)
                #Company Name
                try:
                    Cname = driver.find_element(By.XPATH, '//*[@id="section-to-print"]/h2[1]/i').text
                except :
                    pass
                #Trading Code
                try:
                    CTradingcode = driver.find_element(By.XPATH,
                                                    '//body[1]/div[2]/section[1]/div[1]/div[3]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/th[1]').text
                    CTradingcode = CTradingcode.strip("Trading Code:")
                except :
                    pass
                #Opening Price
                try:
                    OpeningPrice = driver.find_element(By.XPATH,
                                                    '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[5]/td[1]').text
                except :
                    pass
                #Last day Closing Price
                try:
                    lastDayClosingPrice = driver.find_element(By.XPATH,
                                                    '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[7]/td[1]').text
                except :
                    pass
                #Last Trading Price
                try:
                    LastTradingPrice = driver.find_element(By.XPATH, '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[1]/td[1]').text
                except :
                    pass
                #Closing Price
                try:
                    ClosingPrice = driver.find_element(By.XPATH, '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[1]/td[2]').text
                except :
                    pass
                #Days Volume
                try:
                    volume = driver.find_element(By.XPATH, '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[5]/td[2]').text
                except :
                    pass
                #Days Trade
                try:
                    trade = driver.find_element(By.XPATH, '/html/body/div[2]/section/div/div[3]/div[1]/div/div[3]/table/tbody/tr[6]/td[2]').text
                except :
                    pass

                #Category
                try:
                    category = driver.find_element(By.XPATH, '/html/body/div[2]/section/div/div[3]/div[1]/div/div[11]/table/tbody/tr[2]/td[2]').text
                except :
                    pass
                # Cash Divident
                try:
                    cashDivident = driver.find_element(By.XPATH, "/html/body/div[2]/section/div/div[3]/div[1]/div/div[5]/table/tbody/tr[1]/td").text
                except:
                    pass
                # Stock Divident
                try:
                    stockDivident = driver.find_element(By.XPATH, "/html/body/div[2]/section/div/div[3]/div[1]/div/div[5]/table/tbody/tr[2]/td").text
                except:
                    pass
                # Debaut Trading Code
                try:
                    debTradingCode = driver.find_element(By.XPATH, "/html/body/div[2]/section/div/div[3]/div[1]/div/div[4]/table/tbody/tr[1]/td[2]").text
                except:
                    pass
                # Share Holder Percentage - Sponsor/Director
                try:
                    sponsor = driver.find_element(By.XPATH, '//*[@id="company"]/tbody/tr[6]/td[2]/table/tbody/tr/td[1]').text
                    sponsor = re.findall(pattern,sponsor)
                    sponsor = sponsor[0]

                except:
                    pass
                # Share Holder Percentage - Govt
                try:
                    govt = driver.find_element(By.XPATH, '//*[@id="company"]/tbody/tr[6]/td[2]/table/tbody/tr/td[2]').text
                    govt = re.findall(pattern, govt)
                    govt = govt[0]
                except:
                    pass
                # Share Holder Percentage - Institute
                try:
                    institute = driver.find_element(By.XPATH, '//*[@id="company"]/tbody/tr[6]/td[2]/table/tbody/tr/td[3]').text
                    institute = re.findall(pattern, institute)
                    institute = institute[0]
                except:
                    pass
                # Share Holder Percentage - Foreign
                try:
                    foreign = driver.find_element(By.XPATH, '//*[@id="company"]/tbody/tr[6]/td[2]/table/tbody/tr/td[4]').text
                    foreign = re.findall(pattern, foreign)
                    foreign = foreign[0]
                except:
                    pass
                # Share Holder Percentage - Public
                try:
                    public = driver.find_element(By.XPATH, '//*[@id="company"]/tbody/tr[6]/td[2]/table/tbody/tr/td[5]').text
                    public = re.findall(pattern, public)
                    public = public[0]
                except:
                    pass

                #Address
                try:
                    Address = driver.find_element(By.CSS_SELECTOR,
                                              'body > div:nth-child(3) > section:nth-child(3) > div:nth-child(1) > div:nth-child(3) > div:nth-child(2) > div:nth-child(1) > div:nth-child(26) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(3)').text
                    
                except :
                    pass
                #Contact
                try:
                    contact = driver.find_element(By.XPATH,
                                              '/html/body/div[2]/section/div/div[3]/div[1]/div/div[13]/table/tbody/tr[3]/td[2]').text

                except :
                    pass
                #Email
                try:
                    email = driver.find_element(By.XPATH,
                                              '/html/body/div[2]/section/div/div[3]/div[1]/div/div[13]/table/tbody/tr[5]/td[2]').text

                except :
                    pass
                #Website
                try:
                    website = driver.find_element(By.XPATH,
                                              '//*[@id="company"]/tbody/tr[6]/td[2]/a').text

                except :
                    pass

                #print(CurlSn, Cname, CTradingcode,OpeningPrice,lastDayClosingPrice, LastTradingPrice, ClosingPrice,volume,trade,category,cashDivident,stockDivident,debTradingCode,sponsor,govt,institute,foreign,public, Address,contact,email, website, CUrl)
                    

                ws.append([CurlSn, Cname, CTradingcode,OpeningPrice,lastDayClosingPrice, LastTradingPrice, ClosingPrice,volume,trade,category,cashDivident,stockDivident,debTradingCode,sponsor,govt,institute,foreign,public, Address,contact,email, website, CUrl])
                

            else:
                break

        wb.save("Result/Companies-data.xlsx")












driver = OpenBrowser.driver
baseUrl = OpenBrowser.baseUrl
object = CompanyDatas()
object.datas()
import OpenBrowser
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time




class companyUrl:
    def url(self):
        driver.get(baseUrl+"company_listing.php")
        time.sleep(10)
        driver.implicitly_wait(3)

        wb = Workbook()
        ws = wb.active
        urls = driver.find_elements(By.CLASS_NAME,"ab1")
        ws['A1'] = "SN"
        ws['B1'] = "URL"
        sn = 1
        for url in urls:
            allUrl = url.get_attribute("href")

            if(allUrl[0:36] == "https://www.dsebd.org/displayCompany"):

                if(allUrl[0:48] == "https://www.dsebd.org/displayCompany.php?name=TB"):
                    print("This Company is not integrated:" + allUrl)
                else:
                    ws.append([sn, allUrl])
                    sn += 1
            else:
                print("This  is not Company url:" + allUrl)

        wb.save("Result/Companies.xlsx")




driver = OpenBrowser.driver
baseUrl = OpenBrowser.baseUrl
object = companyUrl()
object.url()
